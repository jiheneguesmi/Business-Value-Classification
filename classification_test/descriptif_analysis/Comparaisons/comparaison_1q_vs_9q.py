"""
Comparaison des deux approches de classification + graphiques
--------------------------------------------------------------
S1 : recapitulatif_complet_20260517_074344.xlsx  — 9 questions — feuille Toutes_Phrases
     colonnes : Phrase, Category
S2 : recapitulatif_complet_20260605_133153.xlsx  — 1 question  — feuille Toutes_Phrases
     colonnes : Phrase, Classification

Jointure   : texte normalisé de la phrase
Label S1   : "Description" → Descriptif | ROI/Notoriété/Obligation → Non-descriptif
Label S2   : "Descriptif"  → Descriptif | "Non-descriptif" → Non-descriptif

Sorties :
  - 4 graphiques PNG dans OUTPUT_DIR
  - Excel  : Detail_Phrases / Metriques_Globales / Detail_Desaccords / Graphiques
"""

import io
import time
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

EXCEL_SCRIPT1 = r"F:\Jihene\business_value_classification\Business-Value-Knowledge-Graph\classification_test\Classification\resultats_classification_phrases\recapitulatif_complet_20260517_074344.xlsx"
EXCEL_SCRIPT2 = r"F:\Jihene\business_value_classification\Business-Value-Knowledge-Graph\classification_test\descriptif_analysis\Classification\approche_1_question\resultats_classification_descriptif_1question\recapitulatif_complet_20260605_133153.xlsx"
OUTPUT_DIR    = r"F:\Jihene\business_value_classification\Business-Value-Knowledge-Graph\classification_test\comparaison_approches"

SHEET_NAME = "Toutes_Phrases"

# Palettes
C_DESC    = "#4C9BE8"   # bleu  — Descriptif
C_NONDESC = "#E87B4C"   # orange — Non-descriptif
C_AGREE   = "#5DBB6B"   # vert  — accord
C_DISAG   = "#E85C5C"   # rouge — désaccord
C_S1      = "#2D6A9F"   # bleu foncé — S1
C_S2      = "#E8A84C"   # ambre      — S2

# ═══════════════════════════════════════════════════════════════════════════════


# ── Helpers données ────────────────────────────────────────────────────────────

def normalize(text) -> str:
    return " ".join(str(text).strip().split())

def to_binary_s1(label: str) -> str:
    return "Descriptif" if str(label).strip().lower() == "description" else "Non-descriptif"

def to_binary_s2(label: str) -> str:
    return "Descriptif" if str(label).strip().lower() == "descriptif" else "Non-descriptif"


# ── Helpers styling Excel ──────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="2D6A9F", end_color="2D6A9F")
AGREE_FILL    = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
DISAGREE_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
SUBHEAD_FILL  = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHEAD_FONT  = Font(bold=True, name="Arial", size=10)
BODY_FONT     = Font(name="Arial", size=10)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN

def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 65)


# ── Helpers graphiques ─────────────────────────────────────────────────────────

def _savefig_bytes(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf.read()

def _bar_labels(ax, fmt="{:.1f}%", fontsize=10):
    for p in ax.patches:
        h = p.get_height()
        if h > 0:
            ax.annotate(
                fmt.format(h),
                (p.get_x() + p.get_width() / 2, h),
                ha="center", va="bottom", fontsize=fontsize, fontweight="bold",
            )

def _pie_labels(fmt="{:.1f}%"):
    return lambda pct: fmt.format(pct) if pct > 2 else ""


# ══════════════════════════════════════════════════════════════════════════════
# GRAPHIQUE 1 — Distribution globale : qui dit plus Descriptif ?
# ══════════════════════════════════════════════════════════════════════════════

def plot_distribution(pct_desc_s1, pct_nondesc_s1, pct_desc_s2, pct_nondesc_s2) -> bytes:
    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor("#F8F9FA")
    ax.set_facecolor("#F8F9FA")

    x      = np.arange(2)
    width  = 0.35
    labels = ["S1\n(9 questions)", "S2\n(1 question)"]
    desc_vals    = [pct_desc_s1,    pct_desc_s2]
    nondesc_vals = [pct_nondesc_s1, pct_nondesc_s2]

    b1 = ax.bar(x - width/2, desc_vals,    width, label="Descriptif",     color=C_DESC,    zorder=3)
    b2 = ax.bar(x + width/2, nondesc_vals, width, label="Non-descriptif", color=C_NONDESC, zorder=3)

    _bar_labels(ax)

    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=12)
    ax.set_ylabel("Pourcentage des phrases communes (%)", fontsize=10)
    ax.set_title(
        "Distribution Descriptif / Non-descriptif par script\n"
        "(sur les phrases communes)",
        fontsize=13, fontweight="bold", pad=14,
    )
    ax.set_ylim(0, 110)
    ax.yaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
    ax.set_axisbelow(True)
    ax.legend(fontsize=10)

    # Annotation "qui dit plus descriptif"
    diff = pct_desc_s1 - pct_desc_s2
    if abs(diff) > 0.5:
        plus = "S1 (9 questions)" if diff > 0 else "S2 (1 question)"
        ax.text(
            0.5, 0.97,
            f"→ {plus} classe davantage en Descriptif (+{abs(diff):.1f} pts)",
            transform=ax.transAxes, ha="center", va="top",
            fontsize=10, color="#333333",
            bbox=dict(boxstyle="round,pad=0.4", facecolor="#FFFBCC", edgecolor="#CCAA00", alpha=0.9),
        )

    fig.tight_layout()
    data = _savefig_bytes(fig)
    plt.close(fig)
    return data


# ══════════════════════════════════════════════════════════════════════════════
# GRAPHIQUE 2 — Décomposition des désaccords : qui dit Descriptif quand l'autre ne le fait pas ?
# ══════════════════════════════════════════════════════════════════════════════

def plot_desaccords(n_s1desc_s2nondesc, n_s1nondesc_s2desc, n_total_disag) -> bytes:
    fig, axes = plt.subplots(1, 2, figsize=(11, 5))
    fig.patch.set_facecolor("#F8F9FA")

    # ── Camembert
    ax_pie = axes[0]
    ax_pie.set_facecolor("#F8F9FA")
    sizes  = [n_s1desc_s2nondesc, n_s1nondesc_s2desc]
    colors = [C_S1, C_S2]
    explode = (0.05, 0.05)
    wedges, texts, autotexts = ax_pie.pie(
        sizes, labels=None, colors=colors, explode=explode,
        autopct=_pie_labels(), startangle=90,
        textprops={"fontsize": 12, "fontweight": "bold"},
        wedgeprops={"edgecolor": "white", "linewidth": 2},
    )
    ax_pie.set_title(
        f"Répartition des {n_total_disag} désaccords",
        fontsize=12, fontweight="bold",
    )
    legend_labels = [
        f"S1 dit Descriptif, S2 dit Non-desc  (n={n_s1desc_s2nondesc})",
        f"S2 dit Descriptif, S1 dit Non-desc  (n={n_s1nondesc_s2desc})",
    ]
    ax_pie.legend(
        wedges, legend_labels,
        loc="lower center", bbox_to_anchor=(0.5, -0.22),
        fontsize=9, framealpha=0.8,
    )

    # ── Barres horizontales
    ax_bar = axes[1]
    ax_bar.set_facecolor("#F8F9FA")
    categories = [
        "S1 (9q) dit Descriptif\nS2 (1q) dit Non-desc",
        "S2 (1q) dit Descriptif\nS1 (9q) dit Non-desc",
    ]
    values = [n_s1desc_s2nondesc, n_s1nondesc_s2desc]
    pcts   = [v / n_total_disag * 100 if n_total_disag else 0 for v in values]
    bars   = ax_bar.barh(categories, values, color=[C_S1, C_S2], edgecolor="white", height=0.5, zorder=3)
    ax_bar.xaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
    ax_bar.set_axisbelow(True)
    for bar, val, pct in zip(bars, values, pcts):
        ax_bar.text(
            bar.get_width() + 0.3, bar.get_y() + bar.get_height() / 2,
            f"{val}  ({pct:.1f}%)",
            va="center", fontsize=10, fontweight="bold",
        )
    ax_bar.set_xlabel("Nombre de phrases", fontsize=10)
    ax_bar.set_title("Qui dit Descriptif quand l'autre dit Non-desc ?", fontsize=12, fontweight="bold")
    ax_bar.set_xlim(0, max(values) * 1.35)

    # Conclusion
    if n_s1desc_s2nondesc > n_s1nondesc_s2desc:
        concl = "S1 (9 questions) est plus permissif\n→ classe plus en Descriptif lors des désaccords"
        color = C_S1
    elif n_s1nondesc_s2desc > n_s1desc_s2nondesc:
        concl = "S2 (1 question) est plus permissif\n→ classe plus en Descriptif lors des désaccords"
        color = C_S2
    else:
        concl = "Les deux scripts sont équilibrés\ndans les désaccords"
        color = "#555555"

    fig.text(
        0.5, -0.04, concl,
        ha="center", fontsize=11, fontweight="bold", color=color,
        bbox=dict(boxstyle="round,pad=0.5", facecolor="#FFFBCC", edgecolor="#CCAA00", alpha=0.95),
    )

    fig.suptitle("Analyse des désaccords entre les deux scripts", fontsize=14, fontweight="bold", y=1.02)
    fig.tight_layout()
    data = _savefig_bytes(fig)
    plt.close(fig)
    return data


# ══════════════════════════════════════════════════════════════════════════════
# GRAPHIQUE 3 — Accord / Désaccord global : Descriptif les deux / Non-desc les deux / Désaccord
# ══════════════════════════════════════════════════════════════════════════════

def plot_accord_desaccord(n_dd, n_nn, n_disag, n_total) -> bytes:
    """
    Donut  (gauche)  : 3 parts — Descriptif les deux / Non-desc les deux / Désaccord
    Barres (droite)  : même données en barres horizontales avec effectifs + %
    """
    pct_dd    = n_dd    / n_total * 100
    pct_nn    = n_nn    / n_total * 100
    pct_disag = n_disag / n_total * 100

    labels  = ["Descriptif\ndans les deux", "Non-descriptif\ndans les deux", "Désaccord"]
    values  = [n_dd, n_nn, n_disag]
    colors  = [C_DESC, C_NONDESC, C_DISAG]
    pcts    = [pct_dd, pct_nn, pct_disag]

    fig, (ax_donut, ax_bar) = plt.subplots(1, 2, figsize=(12, 5))
    fig.patch.set_facecolor("#F8F9FA")

    # ── Donut ──────────────────────────────────────────────────────────────────
    ax_donut.set_facecolor("#F8F9FA")
    wedge_props = {"edgecolor": "white", "linewidth": 3}
    wedges, _, autotexts = ax_donut.pie(
        values,
        labels=None,
        colors=colors,
        autopct=lambda p: f"{p:.1f}%" if p > 2 else "",
        startangle=90,
        pctdistance=0.72,
        wedgeprops=wedge_props,
        textprops={"fontsize": 11, "fontweight": "bold", "color": "white"},
    )
    # Trou du donut
    centre_circle = plt.Circle((0, 0), 0.52, fc="#F8F9FA")
    ax_donut.add_patch(centre_circle)
    # Texte central
    ax_donut.text(
        0, 0.12, f"{n_total}", ha="center", va="center",
        fontsize=22, fontweight="bold", color="#333333",
    )
    ax_donut.text(
        0, -0.18, "phrases\ncommunes", ha="center", va="center",
        fontsize=10, color="#666666",
    )
    ax_donut.legend(
        wedges, [f"{l.replace(chr(10), ' ')}  (n={v})" for l, v in zip(labels, values)],
        loc="lower center", bbox_to_anchor=(0.5, -0.18),
        fontsize=9, framealpha=0.8,
    )
    ax_donut.set_title("Répartition globale\nAccord / Désaccord", fontsize=12, fontweight="bold")

    # ── Barres horizontales ────────────────────────────────────────────────────
    ax_bar.set_facecolor("#F8F9FA")
    y_pos = np.arange(len(labels))
    bars  = ax_bar.barh(
        y_pos, values,
        color=colors, edgecolor="white", linewidth=2,
        height=0.5, zorder=3,
    )
    ax_bar.xaxis.grid(True, linestyle="--", alpha=0.4, zorder=0)
    ax_bar.set_axisbelow(True)
    ax_bar.set_yticks(y_pos)
    ax_bar.set_yticklabels([l.replace("\n", " ") for l in labels], fontsize=11)
    ax_bar.set_xlabel("Nombre de phrases", fontsize=10)
    ax_bar.set_title("Effectifs et pourcentages", fontsize=12, fontweight="bold")
    ax_bar.set_xlim(0, max(values) * 1.38)

    for bar, val, pct in zip(bars, values, pcts):
        ax_bar.text(
            bar.get_width() + max(values) * 0.02,
            bar.get_y() + bar.get_height() / 2,
            f"{val}  ({pct:.1f}%)",
            va="center", fontsize=11, fontweight="bold", color="#333333",
        )

    # Ligne d'accord total
    taux_accord = (n_dd + n_nn) / n_total * 100
    fig.text(
        0.5, -0.04,
        f"Taux d'accord global : {taux_accord:.1f}%   |   Taux de désaccord : {pct_disag:.1f}%",
        ha="center", fontsize=12, fontweight="bold", color="#333333",
        bbox=dict(boxstyle="round,pad=0.45", facecolor="#FFFBCC", edgecolor="#CCAA00", alpha=0.95),
    )

    fig.suptitle(
        "Accord entre S1 (9 questions) et S2 (1 question)",
        fontsize=14, fontweight="bold", y=1.02,
    )
    fig.tight_layout()
    data = _savefig_bytes(fig)
    plt.close(fig)
    return data


# ══════════════════════════════════════════════════════════════════════════════
# GRAPHIQUE 4 — Matrice de confusion 2x2
# ══════════════════════════════════════════════════════════════════════════════

def plot_confusion(n_dd, n_nn, n_dn, n_nd, n_total) -> bytes:
    fig, ax = plt.subplots(figsize=(7, 5.5))
    fig.patch.set_facecolor("#F8F9FA")
    ax.set_facecolor("#F8F9FA")
    ax.axis("off")

    matrix = np.array([[n_dd, n_dn], [n_nd, n_nn]])
    pcts   = matrix / n_total * 100

    cell_colors = [
        [C_AGREE,   C_DISAG],
        [C_DISAG,   C_AGREE],
    ]
    col_labels = ["S2 : Descriptif", "S2 : Non-descriptif"]
    row_labels = ["S1 : Descriptif", "S1 : Non-descriptif"]

    cell_size = 0.28
    x0, y0    = 0.18, 0.12

    for i in range(2):
        for j in range(2):
            xpos = x0 + j * (cell_size + 0.04)
            ypos = y0 + (1 - i) * (cell_size + 0.04)
            rect = mpatches.FancyBboxPatch(
                (xpos, ypos), cell_size, cell_size,
                boxstyle="round,pad=0.01",
                facecolor=cell_colors[i][j], edgecolor="white", linewidth=3,
                transform=ax.transAxes, clip_on=False,
            )
            ax.add_patch(rect)
            cx, cy = xpos + cell_size / 2, ypos + cell_size / 2
            n_val  = matrix[i, j]
            pct_val = pcts[i, j]
            ax.text(cx, cy + 0.04, f"{int(n_val)}", transform=ax.transAxes,
                    ha="center", va="center", fontsize=20, fontweight="bold", color="white")
            ax.text(cx, cy - 0.05, f"({pct_val:.1f}%)", transform=ax.transAxes,
                    ha="center", va="center", fontsize=11, color="white")

    # En-têtes colonnes
    for j, label in enumerate(col_labels):
        xpos = x0 + j * (cell_size + 0.04) + cell_size / 2
        ax.text(xpos, y0 + 2 * (cell_size + 0.04) + 0.04, label,
                transform=ax.transAxes, ha="center", va="bottom",
                fontsize=11, fontweight="bold", color=C_S2)

    # En-têtes lignes
    for i, label in enumerate(row_labels):
        ypos = y0 + (1 - i) * (cell_size + 0.04) + cell_size / 2
        ax.text(x0 - 0.03, ypos, label,
                transform=ax.transAxes, ha="right", va="center",
                fontsize=11, fontweight="bold", color=C_S1, rotation=0)

    # Légende accord/désaccord
    patch_acc  = mpatches.Patch(color=C_AGREE,  label="Accord")
    patch_dis  = mpatches.Patch(color=C_DISAG,  label="Désaccord")
    ax.legend(handles=[patch_acc, patch_dis], loc="lower right",
              bbox_to_anchor=(1.0, 0.0), fontsize=10, framealpha=0.8)

    ax.set_title(
        f"Matrice de confusion — {n_total} phrases communes\n"
        "S1 (9 questions)  vs  S2 (1 question)",
        fontsize=13, fontweight="bold", pad=20,
    )
    fig.tight_layout()
    data = _savefig_bytes(fig)
    plt.close(fig)
    return data


# ══════════════════════════════════════════════════════════════════════════════
# GRAPHIQUE 5 — Accord total (accord + désaccord) — jauge / demi-donut
# ══════════════════════════════════════════════════════════════════════════════

def plot_accord_total(n_accord, n_disag, n_total) -> bytes:
    """
    Demi-donut (jauge) centré sur le taux d'accord global
    Accord = Descriptif dans les deux + Non-descriptif dans les deux
    Désaccord = tout le reste
    """
    pct_accord = n_accord / n_total * 100
    pct_disag  = n_disag  / n_total * 100

    fig, ax = plt.subplots(figsize=(7, 4.5))
    fig.patch.set_facecolor("#F8F9FA")
    ax.set_facecolor("#F8F9FA")
    ax.set_aspect("equal")
    ax.axis("off")

    sizes  = [pct_accord, pct_disag]
    colors = [C_AGREE, C_DISAG]

    wedges, _ = ax.pie(
        sizes,
        colors=colors,
        startangle=90,
        counterclock=False,
        wedgeprops={
            "edgecolor": "white",
            "linewidth": 3,
            "width": 0.42
        },
    )

    # Texte central
    ax.text(0,  0.08, f"{pct_accord:.1f}%", ha="center", va="center",
            fontsize=32, fontweight="bold", color=C_AGREE)
    ax.text(0, -0.18, "d'accord", ha="center", va="center",
            fontsize=13, color="#555555")
    ax.text(0, -0.38, f"{n_accord} / {n_total} phrases", ha="center", va="center",
            fontsize=10, color="#777777")

    # Légende
    patch_acc = mpatches.Patch(color=C_AGREE,  label=f"Accord      — {n_accord} phrases  ({pct_accord:.1f}%)")
    patch_dis = mpatches.Patch(color=C_DISAG,  label=f"Désaccord — {n_disag} phrases  ({pct_disag:.1f}%)")
    ax.legend(
        handles=[patch_acc, patch_dis],
        loc="lower center", bbox_to_anchor=(0, -0.08),
        fontsize=10, framealpha=0.85,
    )

    ax.set_title(
        "Taux d'accord global entre S1 (9 questions) et S2 (1 question)",
        fontsize=13, fontweight="bold", pad=10,
    )

    fig.tight_layout()
    data = _savefig_bytes(fig)
    plt.close(fig)
    return data


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def run():
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    print("=" * 70)
    print("COMPARAISON DES DEUX APPROCHES — avec graphiques")
    print("=" * 70)

    # ── Chargement ─────────────────────────────────────────────────────────────
    print("\n[1/5] Chargement des données …")
    df1 = pd.read_excel(EXCEL_SCRIPT1, sheet_name=SHEET_NAME, dtype=str)
    df2 = pd.read_excel(EXCEL_SCRIPT2, sheet_name=SHEET_NAME, dtype=str)
    print(f"  S1 : {len(df1)} lignes  |  S2 : {len(df2)} lignes")

    for name, df, col_phrase, col_label in [
        ("Script 1", df1, "Phrase", "Category"),
        ("Script 2", df2, "Phrase", "Classification"),
    ]:
        for col in [col_phrase, col_label]:
            if col not in df.columns:
                raise ValueError(f"Colonne '{col}' introuvable dans {name}. Dispo : {list(df.columns)}")

    df1 = df1[["Phrase", "Category"]].copy()
    df1.columns = ["phrase_raw", "label_raw"]
    df1["phrase_norm"] = df1["phrase_raw"].apply(normalize)
    df1["label_bin"]   = df1["label_raw"].apply(to_binary_s1)

    df2 = df2[["Phrase", "Classification"]].copy()
    df2.columns = ["phrase_raw", "label_raw"]
    df2["phrase_norm"] = df2["phrase_raw"].apply(normalize)
    df2["label_bin"]   = df2["label_raw"].apply(to_binary_s2)

    df1_d = df1.drop_duplicates("phrase_norm", keep="first")
    df2_d = df2.drop_duplicates("phrase_norm", keep="first")

    # ── Jointure ───────────────────────────────────────────────────────────────
    print("[2/5] Jointure …")
    merged = pd.merge(
        df1_d[["phrase_norm", "phrase_raw", "label_raw", "label_bin"]],
        df2_d[["phrase_norm", "label_raw", "label_bin"]],
        on="phrase_norm", suffixes=("_s1", "_s2"),
    )
    n = len(merged)
    if n == 0:
        print("  ERREUR : aucune phrase commune.")
        return
    print(f"  {n} phrases communes  (non communes S1:{len(df1_d)-n}  S2:{len(df2_d)-n})")

    # ── Métriques ──────────────────────────────────────────────────────────────
    print("[3/5] Calcul des métriques …")

    merged["accord"] = merged["label_bin_s1"] == merged["label_bin_s2"]
    merged["cas"] = merged.apply(
        lambda r: (
            "Descriptif dans les deux"
            if r["label_bin_s1"] == "Descriptif" and r["label_bin_s2"] == "Descriptif"
            else "Non-descriptif dans les deux"
            if r["label_bin_s1"] == "Non-descriptif" and r["label_bin_s2"] == "Non-descriptif"
            else "Désaccord : S1=Desc  / S2=NonDesc"
            if r["label_bin_s1"] == "Descriptif"
            else "Désaccord : S1=NonDesc / S2=Desc"
        ), axis=1,
    )

    taux_accord    = merged["accord"].mean() * 100
    taux_desaccord = 100 - taux_accord

    n_desc2    = (merged["cas"] == "Descriptif dans les deux").sum()
    n_nondesc2 = (merged["cas"] == "Non-descriptif dans les deux").sum()
    n_disag    = (~merged["accord"]).sum()

    pct_desc2    = n_desc2    / n * 100
    pct_nondesc2 = n_nondesc2 / n * 100
    pct_disag    = n_disag    / n * 100

    pct_desc_s1    = (merged["label_bin_s1"] == "Descriptif").mean()     * 100
    pct_nondesc_s1 = (merged["label_bin_s1"] == "Non-descriptif").mean() * 100
    pct_desc_s2    = (merged["label_bin_s2"] == "Descriptif").mean()     * 100
    pct_nondesc_s2 = (merged["label_bin_s2"] == "Non-descriptif").mean() * 100

    # Désaccords détaillés
    n_s1desc_s2nondesc = (merged["cas"] == "Désaccord : S1=Desc  / S2=NonDesc").sum()
    n_s1nondesc_s2desc = (merged["cas"] == "Désaccord : S1=NonDesc / S2=Desc").sum()

    print("\n" + "─" * 70)
    print(f"  Phrases communes          : {n}")
    print(f"  Taux d'accord global      : {taux_accord:.1f}%")
    print(f"  Taux de désaccord         : {taux_desaccord:.1f}%")
    print(f"  Descriptif dans les deux  : {n_desc2} ({pct_desc2:.1f}%)")
    print(f"  Non-desc dans les deux    : {n_nondesc2} ({pct_nondesc2:.1f}%)")
    print(f"  Désaccord total           : {n_disag} ({pct_disag:.1f}%)")
    print(f"    dont S1=Desc  / S2=NonDesc : {n_s1desc_s2nondesc}")
    print(f"    dont S1=NonDesc / S2=Desc  : {n_s1nondesc_s2desc}")
    print(f"  S1 → Descriptif:{pct_desc_s1:.1f}%  Non-desc:{pct_nondesc_s1:.1f}%")
    print(f"  S2 → Descriptif:{pct_desc_s2:.1f}%  Non-desc:{pct_nondesc_s2:.1f}%")
    print("─" * 70)

    # ── Graphiques ─────────────────────────────────────────────────────────────
    print("\n[4/5] Génération des graphiques …")
    ts = time.strftime("%Y%m%d_%H%M%S")

    png1 = plot_distribution(pct_desc_s1, pct_nondesc_s1, pct_desc_s2, pct_nondesc_s2)
    png2 = plot_desaccords(n_s1desc_s2nondesc, n_s1nondesc_s2desc, int(n_disag))
    png3 = plot_accord_desaccord(int(n_desc2), int(n_nondesc2), int(n_disag), n)
    png4 = plot_confusion(
        n_dd=int(n_desc2), n_nn=int(n_nondesc2),
        n_dn=int(n_s1desc_s2nondesc), n_nd=int(n_s1nondesc_s2desc),
        n_total=n,
    )

    png5 = plot_accord_total(int(n_desc2 + n_nondesc2), int(n_disag), n)

    # Sauvegarde PNG
    for name, data in [
        (f"graph1_distribution_{ts}.png",   png1),
        (f"graph2_desaccords_{ts}.png",      png2),
        (f"graph3_accord_desaccord_{ts}.png",     png3),
        (f"graph4_confusion_{ts}.png",             png4),
        (f"graph5_accord_total_{ts}.png",           png5),
    ]:
        path = Path(OUTPUT_DIR) / name
        path.write_bytes(data)
        print(f"  PNG → {name}")

    # ── Export Excel ───────────────────────────────────────────────────────────
    print("\n[5/5] Export Excel …")

    df_detail = merged[[
        "phrase_raw", "label_raw_s1", "label_bin_s1",
        "label_raw_s2", "label_bin_s2", "accord", "cas",
    ]].copy()
    df_detail.columns = [
        "Phrase", "Categorie_S1_brute", "Label_S1",
        "Classification_S2_brute", "Label_S2", "Accord", "Cas",
    ]

    df_global = pd.DataFrame([
        {"Métrique": "Phrases communes analysées",       "Valeur": n,                       "Unité": "phrases"},
        {"Métrique": "Taux d'accord global",             "Valeur": round(taux_accord, 2),   "Unité": "%"},
        {"Métrique": "Taux de désaccord global",         "Valeur": round(taux_desaccord, 2),"Unité": "%"},
        {"Métrique": "── Distribution ──",               "Valeur": "",                      "Unité": ""},
        {"Métrique": "Descriptif dans les deux (N)",     "Valeur": int(n_desc2),            "Unité": "phrases"},
        {"Métrique": "Descriptif dans les deux (%)",     "Valeur": round(pct_desc2, 2),     "Unité": "%"},
        {"Métrique": "Non-descriptif dans les deux (N)", "Valeur": int(n_nondesc2),         "Unité": "phrases"},
        {"Métrique": "Non-descriptif dans les deux (%)", "Valeur": round(pct_nondesc2, 2),  "Unité": "%"},
        {"Métrique": "Désaccord total (N)",              "Valeur": int(n_disag),            "Unité": "phrases"},
        {"Métrique": "Désaccord total (%)",              "Valeur": round(pct_disag, 2),     "Unité": "%"},
        {"Métrique": "── Désaccords détaillés ──",       "Valeur": "",                      "Unité": ""},
        {"Métrique": "S1 dit Desc / S2 dit Non-desc (N)","Valeur": int(n_s1desc_s2nondesc),"Unité": "phrases"},
        {"Métrique": "S2 dit Desc / S1 dit Non-desc (N)","Valeur": int(n_s1nondesc_s2desc),"Unité": "phrases"},
        {"Métrique": "── Distribution par script ──",    "Valeur": "",                      "Unité": ""},
        {"Métrique": "S1 (9q) — Descriptif (%)",         "Valeur": round(pct_desc_s1, 2),   "Unité": "%"},
        {"Métrique": "S1 (9q) — Non-descriptif (%)",     "Valeur": round(pct_nondesc_s1, 2),"Unité": "%"},
        {"Métrique": "S2 (1q) — Descriptif (%)",         "Valeur": round(pct_desc_s2, 2),   "Unité": "%"},
        {"Métrique": "S2 (1q) — Non-descriptif (%)",     "Valeur": round(pct_nondesc_s2, 2),"Unité": "%"},
    ])

    df_disag = df_detail[df_detail["Accord"] == False].copy()

    excel_path = Path(OUTPUT_DIR) / f"comparaison_approches_{ts}.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_detail.to_excel(writer, sheet_name="Detail_Phrases",    index=False)
        df_global.to_excel(writer, sheet_name="Metriques_Globales", index=False)
        df_disag.to_excel(writer,  sheet_name="Detail_Desaccords",  index=False)
        # Feuille vide pour les graphiques
        pd.DataFrame().to_excel(writer, sheet_name="Graphiques", index=False)

    # Styling + insertion graphiques
    wb = load_workbook(excel_path)

    # Sheet 1
    ws1 = wb["Detail_Phrases"]
    style_header(ws1)
    accord_col = next((c.column for c in ws1[1] if c.value == "Accord"), None)
    for row in ws1.iter_rows(min_row=2):
        av = ws1.cell(row=row[0].row, column=accord_col).value if accord_col else True
        fill = AGREE_FILL if av in (True, "True", 1) else DISAGREE_FILL
        for cell in row:
            cell.font = BODY_FONT; cell.fill = fill; cell.border = THIN
            cell.alignment = Alignment(vertical="center")
    ws1.freeze_panes = "A2"
    autofit(ws1)

    # Sheet 2
    ws2 = wb["Metriques_Globales"]
    style_header(ws2)
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN; cell.font = BODY_FONT
            v = ws2.cell(row=cell.row, column=1).value or ""
            if str(v).startswith("──"):
                cell.fill = SUBHEAD_FILL; cell.font = SUBHEAD_FONT
            cell.alignment = Alignment(
                horizontal="center" if cell.column > 1 else "left", vertical="center"
            )
    autofit(ws2)

    # Sheet 3
    ws3 = wb["Detail_Desaccords"]
    style_header(ws3)
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT; cell.fill = DISAGREE_FILL
            cell.border = THIN; cell.alignment = Alignment(vertical="center")
    ws3.freeze_panes = "A2"
    autofit(ws3)

    # Sheet 4 — Graphiques
    ws4 = wb["Graphiques"]
    ws4.sheet_view.showGridLines = False
    titles = [
        "Graphique 1 — Distribution globale",
        "Graphique 2 — Analyse des désaccords",
        "Graphique 3 — Accord / Désaccord détaillé",
        "Graphique 4 — Matrice de confusion",
        "Graphique 5 — Taux d'accord total (jauge)",
    ]
    positions = ["A1", "M1", "A32", "M32", "A63"]
    pngs = [png1, png2, png3, png4, png5]

    for title, pos, png_data in zip(titles, positions, pngs):
        col_letter = pos[0] if pos[1].isdigit() else pos[:2]
        row_num    = int(pos.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        title_cell = ws4.cell(
            row=row_num, column=ws4[col_letter + str(row_num)].column
            if hasattr(ws4[col_letter + str(row_num)], "column")
            else ord(col_letter) - 64
        )

        # Titre au-dessus de l'image
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter) if len(col_letter) <= 2 else 1
        tc = ws4.cell(row=row_num, column=col_idx)
        tc.value = title
        tc.font  = Font(bold=True, name="Arial", size=11, color="2D6A9F")

        img = XLImage(io.BytesIO(png_data))
        img.width  = 520
        img.height = 320
        # Placer l'image une ligne en dessous du titre
        from openpyxl.utils.cell import get_column_letter as gcl
        anchor = gcl(col_idx) + str(row_num + 1)
        ws4.add_image(img, anchor)

    wb.save(excel_path)
    print(f"\n  Export Excel → {excel_path.name}")
    print("=" * 70)


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\nInterruption utilisateur")
    except Exception as e:
        import traceback
        print(f"\nErreur : {e}")
        traceback.print_exc()